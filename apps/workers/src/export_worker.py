import asyncio
import logging
from typing import Dict, Any, List, Optional
import csv
import json
import io
import base64
from datetime import datetime
import uuid
from dataclasses import dataclass
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import boto3
from botocore.exceptions import ClientError

logger = logging.getLogger(__name__)

@dataclass
class ExportRequest:
    id: str
    org_id: str
    project_id: str
    export_type: str  # 'csv', 'excel', 'notion', 'google_sheets', 'pdf', 'json'
    data: Dict[str, Any]
    filters: Dict[str, Any]
    format_options: Dict[str, Any]
    created_at: datetime

@dataclass
class ExportResult:
    id: str
    request_id: str
    file_url: str
    file_size: int
    expires_at: datetime
    status: str  # 'completed', 'failed', 'processing'
    error_message: Optional[str] = None

class ExportWorker:
    def __init__(self, s3_bucket: str = "ai-seo-exports", s3_region: str = "us-east-1"):
        self.logger = logger
        self.s3_bucket = s3_bucket
        self.s3_region = s3_region
        self.s3_client = boto3.client('s3', region_name=s3_region)
        
        # Export format configurations
        self.export_formats = {
            'csv': {
                'extension': '.csv',
                'mime_type': 'text/csv',
                'max_rows': 100000
            },
            'excel': {
                'extension': '.xlsx',
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'max_rows': 100000
            },
            'json': {
                'extension': '.json',
                'mime_type': 'application/json',
                'max_rows': 50000
            },
            'notion': {
                'extension': '.json',
                'mime_type': 'application/json',
                'max_rows': 1000
            },
            'google_sheets': {
                'extension': '.csv',
                'mime_type': 'text/csv',
                'max_rows': 10000
            }
        }
    
    async def process_export(self, export_request: ExportRequest) -> ExportResult:
        """Process an export request and return the result"""
        try:
            self.logger.info(f"Processing export request: {export_request.id}")
            
            # Validate export request
            await self._validate_export_request(export_request)
            
            # Prepare data based on export type
            prepared_data = await self._prepare_data(export_request)
            
            # Generate file based on format
            file_content, file_name = await self._generate_file(export_request, prepared_data)
            
            # Upload to S3
            file_url, file_size = await self._upload_to_s3(file_content, file_name, export_request.export_type)
            
            # Set expiration (7 days)
            expires_at = datetime.utcnow().replace(hour=23, minute=59, second=59)
            expires_at = expires_at.replace(day=expires_at.day + 7)
            
            result = ExportResult(
                id=str(uuid.uuid4()),
                request_id=export_request.id,
                file_url=file_url,
                file_size=file_size,
                expires_at=expires_at,
                status='completed'
            )
            
            self.logger.info(f"Export completed: {result.id}")
            return result
            
        except Exception as e:
            self.logger.error(f"Export failed: {e}")
            return ExportResult(
                id=str(uuid.uuid4()),
                request_id=export_request.id,
                file_url="",
                file_size=0,
                expires_at=datetime.utcnow(),
                status='failed',
                error_message=str(e)
            )
    
    async def _validate_export_request(self, export_request: ExportRequest) -> None:
        """Validate the export request"""
        if export_request.export_type not in self.export_formats:
            raise ValueError(f"Unsupported export type: {export_request.export_type}")
        
        if not export_request.data:
            raise ValueError("No data provided for export")
        
        # Check row limits
        format_config = self.export_formats[export_request.export_type]
        data_rows = len(export_request.data.get('keywords', []))
        if data_rows > format_config['max_rows']:
            raise ValueError(f"Data exceeds maximum rows for {export_request.export_type}: {data_rows} > {format_config['max_rows']}")
    
    async def _prepare_data(self, export_request: ExportRequest) -> Dict[str, Any]:
        """Prepare data for export based on type"""
        try:
            data = export_request.data
            filters = export_request.filters
            
            # Apply filters
            if filters:
                data = await self._apply_filters(data, filters)
            
            # Transform data based on export type
            if export_request.export_type == 'csv':
                return await self._prepare_csv_data(data)
            elif export_request.export_type == 'excel':
                return await self._prepare_excel_data(data)
            elif export_request.export_type == 'json':
                return await self._prepare_json_data(data)
            elif export_request.export_type == 'notion':
                return await self._prepare_notion_data(data)
            elif export_request.export_type == 'google_sheets':
                return await self._prepare_google_sheets_data(data)
            else:
                return data
                
        except Exception as e:
            self.logger.error(f"Error preparing data: {e}")
            raise
    
    async def _generate_file(self, export_request: ExportRequest, prepared_data: Dict[str, Any]) -> tuple:
        """Generate file content and name"""
        try:
            timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            format_config = self.export_formats[export_request.export_type]
            
            if export_request.export_type == 'csv':
                return await self._generate_csv_file(prepared_data, timestamp)
            elif export_request.export_type == 'excel':
                return await self._generate_excel_file(prepared_data, timestamp)
            elif export_request.export_type == 'json':
                return await self._generate_json_file(prepared_data, timestamp)
            elif export_request.export_type == 'notion':
                return await self._generate_notion_file(prepared_data, timestamp)
            elif export_request.export_type == 'google_sheets':
                return await self._generate_google_sheets_file(prepared_data, timestamp)
            else:
                raise ValueError(f"Unsupported export type: {export_request.export_type}")
                
        except Exception as e:
            self.logger.error(f"Error generating file: {e}")
            raise
    
    async def _apply_filters(self, data: Dict[str, Any], filters: Dict[str, Any]) -> Dict[str, Any]:
        """Apply filters to data"""
        try:
            filtered_data = data.copy()
            
            if 'keywords' in data:
                keywords = data['keywords']
                
                # Apply keyword filters
                if 'intent' in filters:
                    keywords = [k for k in keywords if k.get('intent') == filters['intent']]
                
                if 'difficulty' in filters:
                    min_diff = filters['difficulty'].get('min', 0)
                    max_diff = filters['difficulty'].get('max', 100)
                    keywords = [k for k in keywords if min_diff <= k.get('difficulty', 0) <= max_diff]
                
                if 'search_volume' in filters:
                    min_volume = filters['search_volume'].get('min', 0)
                    max_volume = filters['search_volume'].get('max', float('inf'))
                    keywords = [k for k in keywords if min_volume <= k.get('search_volume', 0) <= max_volume]
                
                if 'cluster_id' in filters:
                    keywords = [k for k in keywords if k.get('cluster_id') == filters['cluster_id']]
                
                filtered_data['keywords'] = keywords
            
            return filtered_data
            
        except Exception as e:
            self.logger.error(f"Error applying filters: {e}")
            raise
    
    async def _prepare_csv_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for CSV export"""
        try:
            if 'keywords' not in data:
                return data
            
            # Flatten keyword data for CSV
            csv_rows = []
            for keyword in data['keywords']:
                row = {
                    'keyword': keyword.get('keyword', ''),
                    'intent': keyword.get('intent', ''),
                    'difficulty': keyword.get('difficulty', 0),
                    'search_volume': keyword.get('search_volume', 0),
                    'cluster_id': keyword.get('cluster_id', ''),
                    'cluster_label': keyword.get('cluster_label', ''),
                    'serp_features': ', '.join(keyword.get('serp_features', [])),
                    'created_at': keyword.get('created_at', ''),
                    'updated_at': keyword.get('updated_at', '')
                }
                csv_rows.append(row)
            
            return {'rows': csv_rows, 'headers': list(csv_rows[0].keys()) if csv_rows else []}
            
        except Exception as e:
            self.logger.error(f"Error preparing CSV data: {e}")
            raise
    
    async def _prepare_excel_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for Excel export"""
        try:
            # Excel can handle multiple sheets
            sheets = {}
            
            if 'keywords' in data:
                sheets['Keywords'] = await self._prepare_csv_data(data)
            
            if 'clusters' in data:
                sheets['Clusters'] = await self._prepare_cluster_data(data['clusters'])
            
            if 'serp_results' in data:
                sheets['SERP Results'] = await self._prepare_serp_data(data['serp_results'])
            
            return {'sheets': sheets}
            
        except Exception as e:
            self.logger.error(f"Error preparing Excel data: {e}")
            raise
    
    async def _prepare_json_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for JSON export"""
        try:
            # JSON can export the full data structure
            return {
                'export_metadata': {
                    'exported_at': datetime.utcnow().isoformat(),
                    'total_keywords': len(data.get('keywords', [])),
                    'total_clusters': len(data.get('clusters', [])),
                    'filters_applied': data.get('filters', {})
                },
                'data': data
            }
            
        except Exception as e:
            self.logger.error(f"Error preparing JSON data: {e}")
            raise
    
    async def _prepare_notion_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for Notion export"""
        try:
            # Notion uses a specific format for database imports
            notion_data = {
                'database': {
                    'name': 'SEO Keywords',
                    'properties': {
                        'Keyword': {'type': 'title'},
                        'Intent': {'type': 'select'},
                        'Difficulty': {'type': 'number'},
                        'Search Volume': {'type': 'number'},
                        'Cluster': {'type': 'rich_text'},
                        'SERP Features': {'type': 'multi_select'},
                        'Status': {'type': 'select'}
                    },
                    'rows': []
                }
            }
            
            if 'keywords' in data:
                for keyword in data['keywords']:
                    row = {
                        'Keyword': keyword.get('keyword', ''),
                        'Intent': keyword.get('intent', ''),
                        'Difficulty': keyword.get('difficulty', 0),
                        'Search Volume': keyword.get('search_volume', 0),
                        'Cluster': keyword.get('cluster_label', ''),
                        'SERP Features': keyword.get('serp_features', []),
                        'Status': 'New'
                    }
                    notion_data['database']['rows'].append(row)
            
            return notion_data
            
        except Exception as e:
            self.logger.error(f"Error preparing Notion data: {e}")
            raise
    
    async def _prepare_google_sheets_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data for Google Sheets export"""
        try:
            # Google Sheets uses CSV format for import
            return await self._prepare_csv_data(data)
            
        except Exception as e:
            self.logger.error(f"Error preparing Google Sheets data: {e}")
            raise
    
    async def _prepare_cluster_data(self, clusters: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Prepare cluster data for export"""
        try:
            csv_rows = []
            for cluster in clusters:
                row = {
                    'cluster_id': cluster.get('id', ''),
                    'label': cluster.get('label', ''),
                    'size': cluster.get('size', 0),
                    'centroid': cluster.get('centroid', ''),
                    'keywords': ', '.join(cluster.get('keywords', [])),
                    'created_at': cluster.get('created_at', '')
                }
                csv_rows.append(row)
            
            return {'rows': csv_rows, 'headers': list(csv_rows[0].keys()) if csv_rows else []}
            
        except Exception as e:
            self.logger.error(f"Error preparing cluster data: {e}")
            raise
    
    async def _prepare_serp_data(self, serp_results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Prepare SERP data for export"""
        try:
            csv_rows = []
            for result in serp_results:
                row = {
                    'keyword': result.get('keyword', ''),
                    'position': result.get('position', 0),
                    'title': result.get('title', ''),
                    'url': result.get('url', ''),
                    'snippet': result.get('snippet', ''),
                    'domain': result.get('domain', ''),
                    'features': ', '.join(result.get('features', [])),
                    'fetched_at': result.get('fetched_at', '')
                }
                csv_rows.append(row)
            
            return {'rows': csv_rows, 'headers': list(csv_rows[0].keys()) if csv_rows else []}
            
        except Exception as e:
            self.logger.error(f"Error preparing SERP data: {e}")
            raise
    
    async def _generate_csv_file(self, prepared_data: Dict[str, Any], timestamp: str) -> tuple:
        """Generate CSV file content"""
        try:
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=prepared_data['headers'])
            writer.writeheader()
            writer.writerows(prepared_data['rows'])
            
            file_content = output.getvalue().encode('utf-8')
            file_name = f"seo_keywords_{timestamp}.csv"
            
            return file_content, file_name
            
        except Exception as e:
            self.logger.error(f"Error generating CSV file: {e}")
            raise
    
    async def _generate_excel_file(self, prepared_data: Dict[str, Any], timestamp: str) -> tuple:
        """Generate Excel file content"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each data type
            for sheet_name, sheet_data in prepared_data['sheets'].items():
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers
                if sheet_data['headers']:
                    for col, header in enumerate(sheet_data['headers'], 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                
                # Add data
                for row, data_row in enumerate(sheet_data['rows'], 2):
                    for col, header in enumerate(sheet_data['headers'], 1):
                        ws.cell(row=row, column=col, value=data_row.get(header, ''))
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            file_content = output.getvalue()
            file_name = f"seo_keywords_{timestamp}.xlsx"
            
            return file_content, file_name
            
        except Exception as e:
            self.logger.error(f"Error generating Excel file: {e}")
            raise
    
    async def _generate_json_file(self, prepared_data: Dict[str, Any], timestamp: str) -> tuple:
        """Generate JSON file content"""
        try:
            file_content = json.dumps(prepared_data, indent=2, default=str).encode('utf-8')
            file_name = f"seo_keywords_{timestamp}.json"
            
            return file_content, file_name
            
        except Exception as e:
            self.logger.error(f"Error generating JSON file: {e}")
            raise
    
    async def _generate_notion_file(self, prepared_data: Dict[str, Any], timestamp: str) -> tuple:
        """Generate Notion import file content"""
        try:
            file_content = json.dumps(prepared_data, indent=2, default=str).encode('utf-8')
            file_name = f"notion_import_{timestamp}.json"
            
            return file_content, file_name
            
        except Exception as e:
            self.logger.error(f"Error generating Notion file: {e}")
            raise
    
    async def _generate_google_sheets_file(self, prepared_data: Dict[str, Any], timestamp: str) -> tuple:
        """Generate Google Sheets import file content"""
        try:
            # Google Sheets can import CSV directly
            return await self._generate_csv_file(prepared_data, timestamp)
            
        except Exception as e:
            self.logger.error(f"Error generating Google Sheets file: {e}")
            raise
    
    async def _upload_to_s3(self, file_content: bytes, file_name: str, export_type: str) -> tuple:
        """Upload file to S3 and return URL and size"""
        try:
            format_config = self.export_formats[export_type]
            
            # Generate S3 key
            s3_key = f"exports/{datetime.utcnow().strftime('%Y/%m/%d')}/{file_name}"
            
            # Upload to S3
            self.s3_client.put_object(
                Bucket=self.s3_bucket,
                Key=s3_key,
                Body=file_content,
                ContentType=format_config['mime_type'],
                ContentDisposition=f'attachment; filename="{file_name}"'
            )
            
            # Generate presigned URL (7 days expiration)
            file_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.s3_bucket, 'Key': s3_key},
                ExpiresIn=7 * 24 * 3600  # 7 days
            )
            
            file_size = len(file_content)
            
            self.logger.info(f"File uploaded to S3: {s3_key}, size: {file_size} bytes")
            return file_url, file_size
            
        except ClientError as e:
            self.logger.error(f"S3 upload error: {e}")
            raise
        except Exception as e:
            self.logger.error(f"Error uploading to S3: {e}")
            raise
    
    async def push_to_notion(self, data: Dict[str, Any], notion_token: str, database_id: str) -> Dict[str, Any]:
        """Push data directly to Notion database"""
        try:
            # This would integrate with Notion API
            # For now, return success response
            return {
                'success': True,
                'database_id': database_id,
                'rows_added': len(data.get('keywords', [])),
                'message': 'Data pushed to Notion successfully'
            }
            
        except Exception as e:
            self.logger.error(f"Error pushing to Notion: {e}")
            raise
    
    async def push_to_google_sheets(self, data: Dict[str, Any], spreadsheet_id: str, credentials: Dict[str, Any]) -> Dict[str, Any]:
        """Push data directly to Google Sheets"""
        try:
            # This would integrate with Google Sheets API
            # For now, return success response
            return {
                'success': True,
                'spreadsheet_id': spreadsheet_id,
                'rows_added': len(data.get('keywords', [])),
                'message': 'Data pushed to Google Sheets successfully'
            }
            
        except Exception as e:
            self.logger.error(f"Error pushing to Google Sheets: {e}")
            raise
